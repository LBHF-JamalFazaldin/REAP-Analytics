import os
import pandas as pd
from openpyxl import load_workbook
import inspect as insp
from sqlalchemy import create_engine
from IPython.display import display as original_display

# -----------------------------
# DATABASE CONNECTION
# -----------------------------

# Database credentials
db_host = 'HF-UKS-LBID01.lbhf.gov.uk'
db_port = '1433'
db_name = 'IA_ODS'

# Create the connection string for SQL Server using pyodbc with Windows Authentication
connection_string = f'mssql+pyodbc://@{db_host}:{db_port}/{db_name}?driver=ODBC+Driver+17+for+SQL+Server&Trusted_Connection=yes'

# Create the database engine
engine = create_engine(connection_string)

# -----------------------------
# UTILITY FUNCTIONS
# -----------------------------

def clean_label(label):
    """
    Cleans a label by replacing underscores with spaces and converting to title case.
    Args:
        label (str): The label string to clean.
    Returns:
        str: The cleaned label.
    """
    try:
        return label.replace('_', ' ').title()
    except AttributeError as e:
        print(f'Error cleaning label: {e}')
        return label
 
def get_var_name(var):
    """
    Attempts to retrieve the name of a variable from the global scope.
 
    Args:
        var (object): The object to look up.
 
    Returns:
        str or None: The variable name if found, else None.
    """
    try:
        for name, value in globals().items():
            if value is var:
                return name
    except Exception as e:
        print(f'Error getting variable name: {e}')
 
 
def header_list(df):
    """
    Extracts the header (first row) of a DataFrame and returns remaining rows as a new DataFrame.
 
    Args:
        df (pd.DataFrame): The DataFrame whose header and body are to be separated.
 
    Returns:
        pd.DataFrame: New DataFrame using first row as header.
    """
    try:
        df_list_ = df.copy()
        df_list = df_list_.columns.tolist()
        df_list = pd.DataFrame(df_list)
        new_header = df_list.iloc[0]
        df_list = df_list[1:]
        df_list.columns = new_header
        df_list.reset_index(drop=True, inplace=True)
        return df_list
    except Exception as e:
        print(f'Error creating header list: {e}')
        return pd.DataFrame()

def strip_dataframe(df):
    """
    Strips leading and trailing whitespace from all string cells and column headers in a DataFrame.
    Args:
        df (pd.DataFrame): The DataFrame to clean.
    Returns:
        pd.DataFrame: A new DataFrame with whitespace stripped from strings and headers.
    """
    # Strip whitespace from column headers
    df.columns = df.columns.str.strip()
 
    # Strip whitespace from string cells
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    return df 
 
def display(df, max_columns=True, max_rows=False, **kwargs):
    """
    Displays a DataFrame with its name and number of records.
    Args:
        df (pd.DataFrame): DataFrame to display.
        max_columns (bool): Show all columns if True.
        max_rows (bool): Show all rows if True.
    """
    try:
        frame = insp.currentframe().f_back
        name = "Unnamed DataFrame"
        for var_name, var_value in frame.f_locals.items():
            if var_value is df:
                name = var_name
                break
 
        if name not in {'df', 'Unnamed DataFrame', 'unique_counts', 'summary'}:
            print(f"DataFrame: {name}")
        if name not in {'info_df', 'unique_df', 'summary', None}:
            number_of_records = df.shape[0]
            number_of_fields = df.shape[1]
            duplicate_count = df.duplicated(keep=False).sum()
            unique_duplicate_count = duplicate_count - df.duplicated(keep='first').sum()
            print(
                f"Number of records: {number_of_records:,}",
                "|",
                f"Number of fields: {number_of_fields:,}"
            )
            print(
                f"Number of unique duplicate records: {unique_duplicate_count}"
                " |",
                f"Total number of duplicate records: {duplicate_count}"
            )
        elif name == 'info_df':
            print(f"Max number of non-null records: {kwargs.get('max_records', 0):,}")
        
        if max_columns:
            pd.set_option('display.max_columns', None)
        if max_rows:
            pd.set_option('display.max_rows', None)
 
        original_display(df)
        pd.reset_option('display.max_columns')
        pd.reset_option('display.max_rows')
 
    except Exception as e:
        print(f'Error displaying DataFrame: {e}')

def read_directory(directory=False, max_files=30):
    """
    Lists all files in a directory. Defaults to the current working directory if none provided.
 
    Args:
        directory (str or bool): Path to the directory or False to use current working directory.
    """
    try:
        if directory == False:
            directory = os.getcwd()
        files = pd.DataFrame(os.listdir(directory))
        files.rename(columns={0:'File Names'}, inplace=True)
        
        if directory == os.getcwd():
            print(f"Your Current Directory is: {directory}")
        else:
            print(f"Directory being read is: {directory}")
            
        if files.shape[0] < max_files:
            display(files, max_rows=True)
        else:
            display(files)
    
    except Exception as e:
        print(f'Error reading directory: {e}')

def unique_values(df, show_df=10, sort_values=True):
    """
    Displays unique values for each column in a DataFrame, with optional sorting.
 
    Args:
        df (pd.DataFrame): The DataFrame to analyze.
        show_df (int): Number of rows to display from the resulting summary DataFrame.
        sort_values (bool): Whether to sort the unique values.
    """
    try:
        unique_df_data = {}
        max_length = 0
 
        for col in df.columns:
            values = df[col].dropna().unique()  # drop NaNs to avoid sorting issues
            if sort_values:
                try:
                    values = sorted(values)  # will work for most types
                except Exception:
                    values = list(values)  # fallback if sorting fails
 
            max_length = max(max_length, len(values))
            unique_df_data[col] = values
 
        # Pad shorter lists with None
        for col in unique_df_data:
            padding = [None] * (max_length - len(unique_df_data[col]))
            unique_df_data[col] += padding
 
        unique_df = pd.DataFrame(unique_df_data)
        unique_df.fillna('--', inplace=True)

        unique_df = unique_df.head(show_df)
 
        if show_df:
            display(unique_df, max_rows=True)

        return
 
    except Exception as e:
        print(f'Error extracting unique values: {e}')
        return
 
 
def validate_data(df, show_df=10):
    """
    Validates a DataFrame by displaying basic metrics such as unique/non-null values,
    data types, duplicates, and descriptive statistics.
 
    Args:
        df (pd.DataFrame): The DataFrame to validate.
    """
    try:
        print('#' * 165, end='')
 
        # Display the dataset
        print('\nValidation Dataframe')
        display(df)

        # Display the unique values of the dataset
        print('Unique Values')
        unique_values(df, show_df=show_df)
        print()
 
        # Unique values and non-null counts
        info_df = pd.DataFrame(df.nunique())
        null_counts = pd.DataFrame(len(df) - df.notnull().sum())
        blank_counts = pd.DataFrame(df.apply(lambda col: col.astype(str).apply(lambda x: x.strip() == '').sum()))
        dtypes = pd.DataFrame(df.dtypes, columns=['Data Type'])
 
        # Merge all metrics into one DataFrame
        info_df = pd.merge(
            info_df, null_counts,
            how='left', left_index=True, right_index=True,
            suffixes=['_unique', '_null']
        )

        info_df = pd.merge(
            info_df, blank_counts,
            how='left', left_index=True, right_index=True
        )
 
        info_df = pd.merge(
            info_df, dtypes,
            how='left', left_index=True, right_index=True
        )
 
        # Format and rename
        info_df.reset_index(inplace=True)
        info_df.rename(
            columns={
                '0_unique': 'No. of Unique Values',
                '0_null': 'No. of Null Values',
                0: 'No. of Blank Values',
                'index': 'Field Name'
            },
            inplace=True
        )

        info_df[['No. of Unique Values', 'No. of Null Values', 'No. of Blank Values']] = info_df[
            ['No. of Unique Values', 'No. of Null Values', 'No. of Blank Values']
        ].map(lambda x: f"{x:,}")
        
        for column in ['No. of Null Values', 'No. of Blank Values']:
            info_df[column] = info_df[column].replace({"0":"--"})
            info_df[column] = info_df[column].apply(lambda x: 'Empty' if x == f"{len(df):,}" else x)

        # Check for duplicates
        duplicate_count = df.duplicated(keep=False).sum()

        display(info_df, max_rows=True, duplicate_count=duplicate_count, max_records=len(df))
 
        # Summary statistics
        print("\nSummary statistics:")
        summary = df.describe()
        display(summary)

        print('End of data validation')
        print('#' * 165)
    except Exception as e:
        print(f'Error validating data: {e}')
 
 
def query_data(schema, data):
    """
    Queries a table from SQL Server and returns the results as a DataFrame.
 
    Args:
        schema (str): Schema name.
        data (str): Table name.
 
    Returns:
        pd.DataFrame: Queried data.
    """
    try:
        query = f'SELECT * FROM [{schema}].[{data}]'
        df = pd.read_sql(query, engine)
        print(f'Successfully imported {data}')
        return df
    except Exception as e:
        print(f'Error querying data: {e}')
        return pd.DataFrame()

def capture_header_format(file_path, sheet_name):
    """
    Capture header cell formatting and column widths from an Excel sheet.
 
    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to inspect.
 
    Returns:
        tuple: (header_styles, col_widths) dictionaries
    """
    header_styles, col_widths = {}, {}
 
    try:
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
 
            # Capture header styles (one cell per column in row 1)
            for cell in ws[1]:
                header_styles[cell.column_letter] = {
                    "font": cell.font.copy(),
                    "fill": cell.fill.copy(),
                    "alignment": cell.alignment.copy(),
                    "border": cell.border.copy(),
                    "number_format": cell.number_format,
                }
 
            # Capture column widths
            for col in ws.iter_cols(1, ws.max_column):
                col_letter = col[0].column_letter
                col_widths[col_letter] = ws.column_dimensions[col_letter].width
 
    except Exception as e:
        print(f"Could not capture formatting: {e}")
 
    return header_styles, col_widths

def apply_header_format(file_path, sheet_name, header_styles, col_widths):
    """
    Apply saved header formatting and column widths to a sheet.
 
    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to format.
        header_styles (dict): Dictionary of cell styles per column letter.
        col_widths (dict): Dictionary of column widths per column letter.
    """
    if not header_styles and not col_widths:
        return  # Nothing to apply
 
    wb = load_workbook(file_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
 
        # Restore header formatting
        for cell in ws[1]:
            if cell.column_letter in header_styles:
                style = header_styles[cell.column_letter]
                cell.font = style["font"]
                cell.fill = style["fill"]
                cell.alignment = style["alignment"]
                cell.border = style["border"]
                cell.number_format = style["number_format"]
 
        # Restore column widths
        for col_letter, width in col_widths.items():
            if width is not None:
                ws.column_dimensions[col_letter].width = float(width)
 
    wb.save(file_path)

def export_file(df, file_type='csv', sheet_name='Sheet1', **kwargs):
    """
    Exports a DataFrame to CSV or XLSX while preserving Excel header formatting if XLSX exists.
 
    Args:
        df (pd.DataFrame): The DataFrame to export.
        file_type (str): 'csv' or 'xlsx' (default: 'csv')
        sheet_name (str): Sheet name for Excel export (default: 'Sheet1')
        **kwargs:
            - directory (str): Target directory path.
            - df_name (str): Optional name for the DataFrame file.
    """
    try:
        directory = kwargs.get(
            'directory',
            r"C:\Users\jf79\OneDrive - Office Shared Service\Documents\H&F Analysis\Python CSV Repositry"
        )
        df_name = kwargs.get('df_name', "exported_dataframe")
        file_type = file_type.lower()
 
        # Prompt for name if not found
        if not isinstance(df_name, str) or df_name.strip() == '_':
            df_name = input('DataFrame not found in global variables. Please enter a file name: ')
 
        # Ensure directory exists
        os.makedirs(directory, exist_ok=True)
 
        file_path = os.path.join(directory, f'{df_name}.{file_type}')
        print(f'Exporting {df_name} to {file_type.upper()}...\n@ {file_path}\n')
 
        if file_type == 'csv':
            df.to_csv(file_path, index=False)
 
        elif file_type == 'xlsx':
            # Step 1: Capture formatting if file exists
            header_styles, col_widths = {}, {}
            if os.path.exists(file_path):
                header_styles, col_widths = capture_header_format(file_path, sheet_name)
 
            # Step 2: Write DataFrame
            mode = "a" if os.path.exists(file_path) else "w"
            with pd.ExcelWriter(file_path, engine="openpyxl", mode=mode, if_sheet_exists="overlay") as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=0)
 
            # Step 3: Reapply formatting
            apply_header_format(file_path, sheet_name, header_styles, col_widths)
 
        else:
            raise ValueError(f"Unsupported file_type '{file_type}'. Use 'csv' or 'xlsx'.")
 
        print(f'Successfully exported {df_name} to {file_type.upper()}')
 
    except Exception as e:
        print(f'Error exporting to {file_type.upper()}: {e}')

def id_check(df, id, keep=False, **kwargs):
    id_check = pd.DataFrame(df[id])

    directory = kwargs.get(
        'directory',
        r"C:\Users\jf79\OneDrive - Office Shared Service\Documents\H&F Analysis\Python CSV Repositry"
    )

    export_file(
        id_check,
        directory=directory,
        df_name='id_check',
        file_type='csv'
    )

    id_check = pd.read_csv('id_check.csv')

    id_check = id_check[id_check[id].duplicated(keep=keep)].sort_values(by=id)
    id_check['count'] = 1

    merge = pd.merge(
        df, id_check,
        how='left', on=id 
    ).dropna(subset=['count'], axis=0).drop(columns='count')

    return merge